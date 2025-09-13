import openpyxl

# Load the workbook and select the active sheet
workbook = openpyxl.load_workbook('example.xlsx')
sheet = workbook.active # You can also select a specific sheet by name
print(workbook.sheetnames)
# sheet = workbook['Sheet1']
# Read data from specific cells
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
    for cell in row:
        print(cell.value, end=" ")
    print()  # New line after each row